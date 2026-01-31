from typing import List
from io import BytesIO

from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font

from src.repositories.result_repository import ResultRepository
from src.repositories.test_repository import TestRepository


class ExportService:
    @staticmethod
    def export_test_results_to_excel(
        db: Session,
        slug: str,
        owner_id: int,
    ) -> BytesIO:
        """
        Создает Excel таблицу с тестами пользователя
        
        :param db: Сессия в бд
        :param slug: Уникальный ID теста
        :param owner_id: ID создателя, кому принадлежат тесты
        :return: BytesIO
        """
        test = TestRepository.get_test_by_slug(db, slug)

        if not test or test.owner_id != owner_id:
            raise ValueError("Нет доступа к тесту")

        results = ResultRepository.get_results_by_test(db, slug)

        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты"

        headers = [
            "№",
            "ФИО",
            "Баллы",
            "Макс. баллы",
            "Дата прохождения",
        ]

        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for idx, r in enumerate(results, start=1):
            ws.append([
                idx,
                r.fullname,
                r.score,
                r.max_score,
                r.finished_at.strftime("%d.%m.%Y %H:%M"),
            ])

        column_widths = [5, 25, 10, 12, 20]
        for i, width in enumerate(column_widths, start=1):
            column_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[column_letter].width = width

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)

        return stream